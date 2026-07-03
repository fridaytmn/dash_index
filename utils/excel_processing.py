from openpyxl.reader.excel import load_workbook
import pandas as pd
import re
import operator

from app import app

operations = {
    "+": operator.add,
    "-": operator.sub,
    "*": operator.mul,
    "/": operator.truediv,
}


def find_cell_by_value(filename, search_value, number_column):  # noqa C901
    """
    Находит ячейку по значению в файле Excel.

    Args:
        filename:  Имя файла Excel.
        search_value: Значение для поиска.

    Returns:
        Кортеж (row, column) с координатами ячейки, если найдена,
        иначе None.
    """
    try:
        workbook = load_workbook(filename, data_only=True)
        for cell in [x for x in workbook.active.columns][number_column]:
            if cell.value == search_value:
                return cell.row, cell.column

        app.server.logger.info(
            f"Значение '{search_value}' не найдено в файле '{filename}'."
        )

    except FileNotFoundError:
        app.server.logger.info(f"Ошибка: Файл '{filename}' не найден.")
        return None
    return None


def update_cell_by_value(filename, row, columns, values):  # noqa C901

    try:
        workbook = load_workbook(filename, data_only=True)
        for i, column in enumerate(columns):
            workbook.active.cell(row=row, column=column).value = values[i]
        workbook.save(filename)
    except TypeError:
        return False
    except FileNotFoundError:
        return False

    return True


def get_value_by_location(filename, row, column):

    try:
        workbook = load_workbook(filename, data_only=True)
    except FileNotFoundError:
        app.server.logger.info(f"Ошибка: Файл '{filename}' не найден.")
        return None

    return workbook.active.cell(row=row, column=column).value


def normalize_article(article):
    if pd.isna(article):
        return ""

    article = str(article).strip()

    return re.sub(r"[^A-Za-zА-Яа-я0-9]", "", article).upper()


def merge_supplier_to_nomenclature(
    df_supplier: pd.DataFrame, df_nomenclature: pd.DataFrame, supplier_name: str
):

    df_nomenclature = df_nomenclature.copy()
    df_supplier = df_supplier.copy()

    df_supplier.iloc[:, 0] = df_supplier.iloc[:, 0].apply(normalize_article)
    df_nomenclature["Supplier_Designation"] = df_nomenclature[
        "Supplier_Designation"
    ].apply(normalize_article)
    df_nomenclature["Поставщик (Supplier)"] = (
        df_nomenclature["Поставщик (Supplier)"].fillna("").astype(str)
    )

    df_nomenclature = df_nomenclature.set_index("Supplier_Designation")
    not_matched_rows = []

    for _, row in df_supplier.iterrows():

        article = row.iloc[0]
        qty = normalize_quantity(row.iloc[1])

        if article in df_nomenclature.index and article != "":
            df_nomenclature.at[article, "Кол-во, шт."] = qty
            df_nomenclature.at[article, "Поставщик (Supplier)"] = (
                f"{supplier_name} ({int(qty) if qty.is_integer() else qty})"
            )

        else:
            not_matched_rows.append(
                {"Артикул": article, "Кол-во": qty, "Поставщик": supplier_name}
            )

    df_not_matched = pd.DataFrame(not_matched_rows)
    df_nomenclature = df_nomenclature.reset_index()

    return df_nomenclature, df_not_matched


def save_nomenclature_to_excel(df: pd.DataFrame, path: str):
    df.to_excel(path, index=False)


def normalize_quantity(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def export_nomenclature_to_store(
    nomenclature_path: str,
    store_path: str,
) -> int:
    """
    Переносит необходимые столбцы из Nomenclature в Store.

    Returns:
        Количество выгруженных строк.
    """

    columns = [
        "Обозначение (Designation)",
        "Характеристики",
        "Главная страница",
        "ПОДРАЗДЕЛ",
        "Область применения P, М, К, N, S, H",
        "Кол-во, шт.",
    ]

    df = pd.read_excel(nomenclature_path)

    missing = [column for column in columns if column not in df.columns]

    if missing:
        raise ValueError(f"В файле отсутствуют столбцы: {', '.join(missing)}")

    result = df[columns].copy()

    result.to_excel(store_path, index=False)

    return len(result)
